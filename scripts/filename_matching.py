from pathlib import Path
import pandas as pd
import openpyxl
import re


def get_df_files(folder: Path | str) -> pd.DataFrame:
    files = []
    for fi in Path(folder).iterdir():
        files.append((fi.stem, fi.suffix.lower(), fi.stem.lower()))
    df = pd.DataFrame(files, columns=["basename", "suffix", "basename_lower"])
    df_pivot = df.pivot(index="basename_lower", columns=["suffix"], values="basename")
    return df_pivot


def replace_extension(file: Path | str, ext_in, ext_out) -> str:
    filename = Path(file).name
    file = Path(file)
    if file.suffix.lower() == ext_in:
        return Path(file).with_suffix(ext_out).as_posix()
    else:
        return file.as_posix()


def match_stem_caseinsensitive(
    filename: str | Path, search_path: str | Path, searched_extension: str
) -> str:
    pattern = re.compile(Path(filename).stem + f"{searched_extension}", re.IGNORECASE)
    files = [f for f in Path(search_path).rglob("*") if pattern.match(f.name)]
    assert len(files) == 1
    return files[0].name


def main():
    folder = Path().joinpath("CTD_data_renamed")
    out = Path().joinpath("CTD_data_renamed/mismatched_filenames/")

    raw_folders = folder.rglob("raw_files")
    out_overview = out.parent.joinpath("overview.xlsx")

    columns = [".xmlcon", ".hex", ".bl", ".mrk", ".hdr", ".btl", ".cnv"]

    content_overview = []
    for di in raw_folders:
        di_name = di.parent.name
        dfi = get_df_files(di)

        relevant_columns = sorted(
            list(set(dfi.columns).intersection(columns)), key=lambda i: columns.index(i)
        )
        additional_columns = set(dfi.columns).difference(relevant_columns)

        output_file = out.joinpath(di_name + "_mismatching_files.xlsx")
        columns_with_mismatch = dfi[relevant_columns].isna().any(axis=1)
        dfi_mismatch = dfi[columns_with_mismatch]
        btl_in_folder = ".btl" in relevant_columns

        if columns_with_mismatch.any() or btl_in_folder:
            sheet_name = "Mismatching filenames"
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                dfi_mismatch.to_excel(
                    writer, sheet_name=sheet_name, startrow=3, index=False
                )

            wb = openpyxl.load_workbook(output_file)
            ws = wb[sheet_name]
            ws["A1"] = f"Addition columns:"
            ws["B1"] = f"{', '.join(additional_columns)}"
            ws["A2"] = f".btl files in raw_files folder:"
            ws["B2"] = f"{'.btl' in relevant_columns}"

            wb.save(output_file)

            content_overview += [
                (
                    di_name,
                    f'=HYPERLINK("{output_file.relative_to(out_overview.parent).as_posix()}", "{di_name}"',
                )
            ]
    df_overview = pd.DataFrame(
        content_overview, columns=["Dataset", "Link"]
    ).sort_values("Dataset")
    df_overview.to_excel(out_overview, index=False)


if __name__ == "__main__":
    main()
